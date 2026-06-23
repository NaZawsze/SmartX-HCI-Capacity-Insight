import io
import os
import tempfile
import unittest
import zipfile
from pathlib import Path
from typing import Optional


try:
    from fastapi.testclient import TestClient
except ModuleNotFoundError:  # pragma: no cover - local host may not have web deps.
    TestClient = None


class FakeReportService:
    def __init__(
        self,
        *,
        month_vms: Optional[list[dict]] = None,
        window_vms: Optional[list[dict]] = None,
        clusters: Optional[list[dict]] = None,
    ) -> None:
        self.month_vms = month_vms
        self.window_vms = window_vms
        self.clusters = clusters

    def latest_report(self, tower_id=None, cluster_id=None, period_days=30, chart_days=365):
        return {
            "scope": {"tower_id": tower_id, "cluster_id": cluster_id},
            "clusters": self.clusters
            if self.clusters is not None
            else [
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A"},
                    "forecast": {
                        "status": "ok",
                        "slope_per_day": 10,
                        "current": 810,
                        "forecast_90d": 1090,
                        "exhaustion_days": None,
                    },
                    "points": [[1764547200, 780], [1764633600, 790], [1764720000, 810]],
                    "total": 1000,
                    "warning": 900,
                },
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-b", "cluster": "Cluster B"},
                    "forecast": {
                        "status": "ok",
                        "slope_per_day": 2,
                        "current": 120,
                        "forecast_90d": 300,
                        "exhaustion_days": 2056,
                    },
                    "points": [[1764547200, 110], [1764633600, 115], [1764720000, 120]],
                    "total": 1000,
                    "warning": 900,
                }
            ],
            "day_fastest_growing_vms": [
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A", "vm_id": "vm-day", "vm": "Day VM"},
                    "forecast": {"status": "ok", "slope_per_day": 4, "current": 240 * 1024**3},
                    "growth_amount": 8 * 1024**3,
                    "previous_value": 232 * 1024**3,
                    "growth_ratio": 0.03,
                }
            ],
            "window_fastest_growing_vms": self.window_vms
            if self.window_vms is not None
            else [
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A", "vm_id": "vm-window-1", "vm": "Window VM One"},
                    "forecast": {"status": "ok", "slope_per_day": 4, "current": 240 * 1024**3},
                    "growth_amount": 20 * 1024**3,
                    "previous_value": 220 * 1024**3,
                    "growth_ratio": 0.09,
                    "window_start_at": "2026-05-22T13:00:00+00:00",
                    "window_end_at": "2026-06-06T19:00:00+00:00",
                },
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A", "vm_id": "vm-window-2", "vm": "Window VM Two"},
                    "forecast": {"status": "ok", "slope_per_day": 3, "current": 140 * 1024**3},
                    "growth_amount": 12 * 1024**3,
                    "previous_value": 128 * 1024**3,
                    "growth_ratio": 0.09,
                    "window_start_at": "2026-05-22T13:00:00+00:00",
                    "window_end_at": "2026-06-06T19:00:00+00:00",
                },
            ],
            "month_fastest_growing_vms": self.month_vms
            if self.month_vms is not None
            else [
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A", "vm_id": "vm-1", "vm": "VM One"},
                    "forecast": {"status": "ok", "slope_per_day": 4, "current": 240 * 1024**3},
                    "growth_amount": 150 * 1024**3,
                    "previous_value": 120 * 1024**3,
                    "growth_ratio": 1.0,
                    "sample_span_days": 31,
                    "window_start_at": "2026-05-01T00:00:00+08:00",
                    "window_end_at": "2026-05-31T00:00:00+08:00",
                }
            ],
            "day_new_vms": [
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A", "vm_id": "vm-new-day", "vm": "New Day VM"},
                    "forecast": {"status": "ok", "slope_per_day": 0, "current": 10 * 1024**3},
                    "first_seen_at": "2026-05-31T09:00:00+08:00",
                }
            ],
            "month_new_vms": [
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-b", "cluster": "Cluster B", "vm_id": "vm-new-month", "vm": "New Month VM"},
                    "forecast": {"status": "ok", "slope_per_day": 0, "current": 30 * 1024**3},
                    "first_seen_at": "2026-05-03T09:00:00+08:00",
                }
            ],
            "cluster_growth_rate": {"per_day": 10, "per_month": 300, "per_quarter": 900},
            "window_days": period_days,
            "chart_days": chart_days,
            "growth_rate_window_days": 7,
            "forecast_days": 90,
            "period_window": {"days": period_days, "start_at": "2026-05-01T00:00:00+08:00", "end_at": "2026-05-31T00:00:00+08:00"},
            "data_window": {"start_at": "2026-05-22T13:00:00+00:00", "end_at": "2026-06-06T19:00:00+00:00"},
            "data_quality": {
                "status": "warning",
                "actual_data_window": {"start_at": "2026-05-22T13:00:00+00:00", "end_at": "2026-06-06T19:00:00+00:00", "days": 16},
                "requested_window": {"days": period_days, "start_at": "2026-05-01T00:00:00+08:00", "end_at": "2026-05-31T00:00:00+08:00"},
                "sample_sufficient": False,
                "missing_collection_dates": ["2026-06-02"],
                "incomplete_clusters": [
                    {
                        "tower_id": 1,
                        "tower": "Tower A",
                        "cluster_id": "cluster-a",
                        "cluster": "Cluster A",
                        "reason": "prometheus_cluster_sample_missing",
                    }
                ],
                "sqlite_vm_count": 2,
                "prometheus_vm_series_count": 1,
                "sqlite_cluster_count": 2,
                "prometheus_cluster_series_count": 1,
                "vm_count_difference": 1,
                "vm_count_difference_ratio": 0.5,
                "latest_collection_status": "partial_failed",
                "latest_success_at": "2026-06-06T19:00:00+00:00",
                "latest_prometheus_sample_at": "2026-06-06T19:00:00+00:00",
                "messages": ["当前报表存在样本不足、缺采或部分集群样本不完整，趋势与预测结论需结合实际采集窗口理解。"],
            },
            "timezone": "Asia/Shanghai",
        }


class V2ReportExportDocumentTest(unittest.TestCase):
    def test_report_period_profiles_cover_all_export_windows(self) -> None:
        from app.v2.reports.export import report_period_profile

        expected = {
            7: ("近 7 天样本增长", "短期突增 VM", "短期（本周）"),
            14: ("近 14 天样本增长", "近两周增长 VM", "短期（两周内）"),
            30: ("近 30 天样本增长", "月度增长 VM", "短期（本月）"),
            90: ("近 90 天样本增长", "季度增长 VM", "短期（本季度）"),
            180: ("近 180 天样本增长", "中长期增长 VM", "短期（半年内）"),
            365: ("近 365 天样本增长", "年度增长 VM", "短期（年度巡检）"),
        }
        for days, (growth_title, vm_title, advice_title) in expected.items():
            with self.subTest(days=days):
                profile = report_period_profile(days)
                self.assertEqual(profile.days, days)
                self.assertEqual(profile.window_growth_label, growth_title)
                self.assertEqual(profile.vm_growth_title, vm_title)
                self.assertEqual(profile.short_advice_title, advice_title)

        self.assertEqual(report_period_profile(999).days, 30)

    def test_docx_and_xlsx_use_period_profile_labels_for_each_window(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx, build_report_xlsx

        expected = {
            7: ("近 7 天样本增长", "短期突增 VM"),
            14: ("近 14 天样本增长", "近两周增长 VM"),
            30: ("近 30 天样本增长", "月度增长 VM"),
            90: ("近 90 天样本增长", "季度增长 VM"),
            180: ("近 180 天样本增长", "中长期增长 VM"),
            365: ("近 365 天样本增长", "年度增长 VM"),
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            for days, (growth_title, vm_title) in expected.items():
                with self.subTest(days=days):
                    report = FakeReportService().latest_report(period_days=days)
                    content, _, _, _ = build_report_docx(report, settings, period_days=days)
                    word_xml, _ = _docx_xml(content)
                    self.assertIn(growth_title, word_xml)
                    self.assertIn(vm_title, word_xml)

                    xlsx_content, _, _, _ = build_report_xlsx(report, settings, period_days=days)
                    workbook_text = "\n".join(_xlsx_values(xlsx_content))
                    self.assertIn(growth_title, workbook_text)
                    self.assertIn(vm_title, workbook_text)

    def test_long_window_with_short_history_uses_profile_sample_notice(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx, build_report_xlsx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            report = FakeReportService().latest_report(period_days=365)
            report["period_window"] = {
                "days": 365,
                "start_at": "2025-06-06T00:00:00+08:00",
                "end_at": "2026-06-06T23:59:59+08:00",
            }
            content, _, _, _ = build_report_docx(report, settings, period_days=365)
            word_xml, footer_xml = _docx_xml(content)
            self.assertIn("近 365 天样本增长", word_xml)
            self.assertIn("采集历史不足对应天数时显示数据不足", word_xml)
            self.assertIn("2026-05-22 - 2026-06-06", word_xml)

            xlsx_content, _, _, _ = build_report_xlsx(report, settings, period_days=365)
            workbook_text = "\n".join(_xlsx_values(xlsx_content))
            self.assertIn("近 365 天样本增长", workbook_text)
            self.assertIn("已按当前可用样本窗口计算", workbook_text)

    def test_docx_uses_customer_template_with_v2_charts_and_vm_count(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            clusters = [
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A"},
                    "forecast": {"status": "ok", "slope_per_day": 2, "current": 120, "forecast_90d": 300, "exhaustion_days": 2056},
                    "points": [[1764547200, 110], [1764633600, 115], [1764720000, 120]],
                    "total": 1000,
                    "warning": 900,
                },
                {
                    "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-b", "cluster": "Cluster B"},
                    "forecast": {"status": "ok", "slope_per_day": 1, "current": 80, "forecast_90d": 160, "exhaustion_days": 3000},
                    "points": [[1764547200, 70], [1764633600, 75], [1764720000, 80]],
                    "total": 1000,
                    "warning": 900,
                },
            ]
            content, _, path, _ = build_report_docx(FakeReportService(clusters=clusters).latest_report(period_days=30), settings, period_days=30)
            self.assertTrue(path.is_file())
            word_xml, footer_xml = _docx_xml(content)
            self.assertIn("存储容量预测平台", word_xml)
            self.assertIn("SMARTX超融合存储容量分析报告", word_xml)
            self.assertIn("SMARTX HCI Storage Capacity Analysis Report", word_xml)
            self.assertIn("客户名称", word_xml)
            self.assertIn("Tower范围", word_xml)
            self.assertIn("集群范围", word_xml)
            self.assertIn("数据质量说明", word_xml)
            self.assertIn("实际采集窗口", word_xml)
            self.assertIn("样本是否足够", word_xml)
            self.assertIn("缺采天数", word_xml)
            self.assertIn("数据不完整集群", word_xml)
            self.assertIn("当前报表存在样本不足、缺采或部分集群样本不完整", word_xml)
            self.assertIn("本报表统计窗口", word_xml)
            self.assertIn("全部集群（2 个）", word_xml)
            self.assertNotIn("SmartX 超融合平台", word_xml)
            self.assertNotIn("存储容量预测分析报告", word_xml)
            self.assertNotIn("Storage Capacity Forecast Report", word_xml)
            self.assertIn("目录", word_xml)
            self.assertIn("一  摘要", word_xml)
            self.assertNotIn("一  执行摘要", word_xml)
            self.assertIn("二  集群容量概览", word_xml)
            self.assertIn("三  集群虚拟机增长分析", word_xml)
            self.assertIn("四  集群容量趋势图表", word_xml)
            self.assertIn("五  全集群汇总报告", word_xml)
            self.assertIn("2.1  范围基本信息", word_xml)
            self.assertIn(">Tower<", word_xml)
            self.assertIn("3.1  Tower A / Cluster A", word_xml)
            self.assertIn("3.2  Tower A / Cluster B", word_xml)
            self.assertIn("4.1  Tower A / Cluster A", word_xml)
            self.assertIn("4.2  Tower A / Cluster B", word_xml)
            self.assertIn("Top 10 VM 增长量", word_xml)
            self.assertIn("增长 VM 样本数", word_xml)
            self.assertIn("当前使用率", word_xml)
            self.assertIn("90 天预测使用率", word_xml)
            self.assertIn("容量阈值", word_xml)
            self.assertIn("容量安全边际", word_xml)
            usage_section = word_xml.split("2.3  容量使用率可视化", 1)[1].split("三  集群虚拟机增长分析", 1)[0]
            self.assertIn("█", usage_section)
            self.assertIn("░", usage_section)
            self.assertIn(f"{'█' * 6}{'░' * 44}", usage_section)
            self.assertIn(f"{'█' * 15}{'░' * 35}", usage_section)
            self.assertNotIn(f"{'█' * 12}{'░' * 38}", usage_section)
            self.assertIn('w:color w:val="1A3C6E"', usage_section)
            self.assertIn('w:color w:val="007ACC"', usage_section)
            self.assertNotIn("w:tbl", usage_section)
            self.assertNotIn("Top 10 集群容量使用率", usage_section)
            self.assertIn("近 14 天样本增长", word_xml)
            self.assertIn("近 30 天样本增长", word_xml)
            self.assertIn("近 90 天样本增长", word_xml)
            self.assertNotIn("近 365 天样本增长", word_xml)
            self.assertIn("采集历史不足对应天数时显示数据不足", word_xml)
            self.assertNotIn("较上月", word_xml)
            self.assertNotIn("较上季度", word_xml)
            self.assertNotIn("较上一年", word_xml)
            self.assertIn('w:fill="1A3C6E"', word_xml)
            self.assertIn('w:fill="F0F4FA"', word_xml)
            self.assertIn('w:color w:val="007ACC"', word_xml)
            self.assertIn("Noto Serif CJK SC", word_xml)
            self.assertNotIn("微软雅黑", word_xml)
            self.assertNotIn("Microsoft YaHei", word_xml)
            self.assertIn("<w:tblHeader", word_xml)
            self.assertIn("<w:cantSplit", word_xml)
            self.assertIn('<w:br w:type="page"/>', word_xml)

            vm_header_row = _docx_row_containing(word_xml, "排名")
            self.assertIn("<w:keepNext", vm_header_row)

            self.assertIn("PAGE", footer_xml)
            self.assertIn("NUMPAGES", footer_xml)
            self.assertIn("第 ", footer_xml)
            self.assertIn(" 页 / 共 ", footer_xml)

            summary_section = word_xml.split("一  摘要", 1)[1].split("二  集群容量概览", 1)[0]
            self.assertEqual(summary_section.count("当前已用容量"), 2)
            self.assertIn("Tower A / Cluster A", summary_section)
            self.assertIn("Tower A / Cluster B", summary_section)

    def test_docx_cover_shows_requested_report_window_separately_from_effective_data_window(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            report = FakeReportService().latest_report(period_days=90)
            report["period_window"] = {
                "days": 90,
                "start_at": "2026-03-11T00:00:00+08:00",
                "end_at": "2026-06-09T23:59:59+08:00",
            }
            report["data_window"] = {
                "start_at": "2026-05-23T00:00:00+08:00",
                "end_at": "2026-06-09T23:59:59+08:00",
            }
            content, _, _, _ = build_report_docx(report, settings, period_days=90)
            word_xml, _ = _docx_xml(content)
            self.assertIn("统计窗口", word_xml)
            self.assertIn("2026-05-23 - 2026-06-09", word_xml)
            self.assertIn("本报表统计窗口", word_xml)
            self.assertIn("近 90 天（2026-03-11 - 2026-06-09）", word_xml)

    def test_docx_directory_precedes_summary_and_lists_cluster_sections(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            content, _, _, _ = build_report_docx(FakeReportService().latest_report(period_days=30), settings, period_days=30)
            word_xml, _ = _docx_xml(content)
            self.assertLess(word_xml.index("目录"), word_xml.index("一  摘要"))
            self.assertIn("3.1  Tower A / Cluster A", word_xml)
            self.assertIn("3.2  Tower A / Cluster B", word_xml)
            self.assertIn("4.1  Tower A / Cluster A", word_xml)
            self.assertIn("4.2  Tower A / Cluster B", word_xml)
            self.assertIn("五  全集群汇总报告", word_xml)

    def test_docx_multi_cluster_summary_orders_risk_tables_and_single_cluster_stays_compact(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        normal_cluster = {
            "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "normal", "cluster": "Normal Cluster"},
            "forecast": {"status": "ok", "current": 40, "forecast_90d": 45, "exhaustion_days": None},
            "points": [[1764547200, 38], [1764633600, 40]],
            "total": 100,
            "warning": 75,
        }
        high_cluster = {
            "labels": {"tower_id": "2", "tower": "Tower B", "cluster_id": "high", "cluster": "High Cluster"},
            "forecast": {"status": "ok", "current": 88, "forecast_90d": 95, "exhaustion_days": 64},
            "points": [[1764547200, 84], [1764633600, 88]],
            "total": 100,
            "warning": 75,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            multi_content, _, _, _ = build_report_docx(
                FakeReportService(clusters=[normal_cluster, high_cluster]).latest_report(period_days=30),
                settings,
                period_days=30,
            )
            multi_xml, _ = _docx_xml(multi_content)
            summary = multi_xml.split("一  摘要", 1)[1].split("二  集群容量概览", 1)[0]
            self.assertEqual(summary.count("当前已用容量"), 2)
            self.assertLess(summary.index("Tower B / High Cluster"), summary.index("Tower A / Normal Cluster"))

            single_content, _, _, _ = build_report_docx(
                FakeReportService(clusters=[normal_cluster]).latest_report(period_days=30),
                settings,
                period_days=30,
            )
            single_xml, _ = _docx_xml(single_content)
            single_summary = single_xml.split("一  摘要", 1)[1].split("二  集群容量概览", 1)[0]
            self.assertEqual(single_summary.count("当前已用容量"), 1)

    def test_docx_uses_native_toc_and_real_heading_styles(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            content, _, _, _ = build_report_docx(FakeReportService().latest_report(period_days=30), settings, period_days=30)
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                word_xml = archive.read("word/document.xml").decode("utf-8")
                settings_xml = archive.read("word/settings.xml").decode("utf-8")

            self.assertIn("TOC", word_xml)
            self.assertIn('1-3', word_xml)
            self.assertIn("请在 Word/WPS 中更新目录", word_xml)
            self.assertIn("<w:updateFields", settings_xml)
            self.assertIn('w:val="true"', settings_xml)
            self.assertNotIn("<w:t>章节</w:t>", word_xml)
            self.assertNotIn("<w:t>内容</w:t>", word_xml)

            for title in ["一  摘要", "二  集群容量概览", "三  集群虚拟机增长分析", "四  集群容量趋势图表", "五  全集群汇总报告"]:
                paragraph_xml = _docx_paragraph_containing(word_xml, title)
                self.assertIn('w:pStyle w:val="Heading1"', paragraph_xml)

            for title in ["3.1  Tower A / Cluster A", "3.2  Tower A / Cluster B", "4.1  Tower A / Cluster A", "4.2  Tower A / Cluster B"]:
                paragraph_xml = _docx_paragraph_containing(word_xml, title)
                self.assertIn('w:pStyle w:val="Heading2"', paragraph_xml)

    def test_docx_cluster_vm_sections_only_include_their_own_cluster_vms(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        window_vms = [
            {
                "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A", "vm_id": "vm-a", "vm": "Cluster A Only VM"},
                "forecast": {"status": "ok", "slope_per_day": 4, "current": 240 * 1024**3},
                "growth_amount": 20 * 1024**3,
                "previous_value": 220 * 1024**3,
                "growth_ratio": 0.09,
                "window_start_at": "2026-05-22T13:00:00+00:00",
                "window_end_at": "2026-06-06T19:00:00+00:00",
            },
            {
                "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-b", "cluster": "Cluster B", "vm_id": "vm-b", "vm": "Cluster B Only VM"},
                "forecast": {"status": "ok", "slope_per_day": 3, "current": 140 * 1024**3},
                "growth_amount": 12 * 1024**3,
                "previous_value": 128 * 1024**3,
                "growth_ratio": 0.09,
                "window_start_at": "2026-05-22T13:00:00+00:00",
                "window_end_at": "2026-06-06T19:00:00+00:00",
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            report = FakeReportService(window_vms=window_vms, month_vms=[]).latest_report(period_days=30)
            content, _, _, _ = build_report_docx(report, settings, period_days=30)
            word_xml, _ = _docx_xml(content)
            cluster_a_section = word_xml.split("3.1  Tower A / Cluster A", 1)[1].split("3.2  Tower A / Cluster B", 1)[0]
            cluster_b_section = word_xml.split("3.2  Tower A / Cluster B", 1)[1].split("四  集群容量趋势图表", 1)[0]
            self.assertIn("Cluster A Only VM", cluster_a_section)
            self.assertNotIn("Cluster B Only VM", cluster_a_section)
            self.assertIn("Cluster B Only VM", cluster_b_section)
            self.assertNotIn("Cluster A Only VM", cluster_b_section)

    def test_docx_cluster_growth_table_uses_fixed_windows_and_data_insufficient_for_short_history(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        base_ts = 1763164800
        gib = 1024**3
        clusters = [
            {
                "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A"},
                "forecast": {"status": "ok", "slope_per_day": 10 * gib, "current": 1140 * gib, "forecast_90d": 1800 * gib},
                "points": [[base_ts, 1000 * gib], [base_ts + 14 * 86_400, 1140 * gib]],
                "total": 2000 * gib,
                "warning": 1800 * gib,
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            report = FakeReportService(clusters=clusters).latest_report(period_days=90)
            content, _, _, _ = build_report_docx(report, settings, period_days=90)
            word_xml, _ = _docx_xml(content)
            self.assertIn("近 14 天样本增长", word_xml)
            self.assertIn("近 30 天样本增长", word_xml)
            self.assertIn("近 90 天样本增长", word_xml)
            self.assertNotIn("近 365 天样本增长", word_xml)
            self.assertIn("140.00 GB", word_xml)
            self.assertIn("数据不足", word_xml)

    def test_docx_scope_information_handles_multiple_towers_and_clusters(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        clusters = [
            {
                "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A"},
                "forecast": {"status": "ok", "slope_per_day": 10, "current": 810, "forecast_90d": 1090},
                "points": [[1764547200, 780], [1764720000, 810]],
                "total": 1000,
                "warning": 900,
            },
            {
                "labels": {"tower_id": "2", "tower": "Tower B", "cluster_id": "cluster-b", "cluster": "Cluster B"},
                "forecast": {"status": "ok", "slope_per_day": 3, "current": 300, "forecast_90d": 570},
                "points": [[1764547200, 260], [1764720000, 300]],
                "total": 1000,
                "warning": 900,
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            content, _, _, _ = build_report_docx(FakeReportService(clusters=clusters).latest_report(period_days=30), settings, period_days=30)
            word_xml, _ = _docx_xml(content)
            self.assertIn("Tower数量", word_xml)
            self.assertIn("2 个", word_xml)
            self.assertIn("集群数量", word_xml)
            self.assertIn("全部 Tower（2 个）", word_xml)
            self.assertIn("全部集群（2 个）", word_xml)
            self.assertIn('<w:t>Tower A</w:t>', word_xml)
            self.assertIn('<w:t>Cluster A</w:t>', word_xml)

    def test_docx_keeps_day_growth_data_when_month_growth_is_empty(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            content, _, _, _ = build_report_docx(FakeReportService(month_vms=[]).latest_report(period_days=7), settings, period_days=7)
            word_xml, _ = _docx_xml(content)
            self.assertIn("虚拟机增长分析", word_xml)
            self.assertIn("Day VM", word_xml)
            self.assertIn("短期突增 VM", word_xml)

    def test_docx_cluster_top100_uses_window_growth_when_month_growth_is_sparse(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            content, _, _, _ = build_report_docx(FakeReportService(month_vms=[]).latest_report(period_days=14), settings, period_days=14)
            word_xml, _ = _docx_xml(content)
            self.assertIn("Window VM One", word_xml)
            self.assertIn("Window VM Two", word_xml)
            self.assertIn("近两周增长 VM", word_xml)

    def test_docx_cluster_top100_merges_window_and_month_growth(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            content, _, _, _ = build_report_docx(FakeReportService().latest_report(period_days=30), settings, period_days=30)
            word_xml, _ = _docx_xml(content)
            self.assertIn("Window VM One", word_xml)
            self.assertIn("Window VM Two", word_xml)
            self.assertIn("VM One", word_xml)
            self.assertIn('w:fill="F7F9FC"', word_xml)
            self.assertIn('w:color w:val="007ACC"', word_xml)
            self.assertIn('w:color w:val="DC3535"', word_xml)

    def test_docx_uses_customer_report_summary_and_rule_based_advice(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx

        chinese_vm = [
            {
                "labels": {
                    "tower_id": "1",
                    "tower": "Tower A",
                    "cluster_id": "cluster-a",
                    "cluster": "Cluster A",
                    "vm_id": "vm-cn",
                    "vm": "边缘云管理平台04",
                },
                "forecast": {"status": "ok", "slope_per_day": 4, "current": 240 * 1024**3},
                "growth_amount": 20 * 1024**3,
                "previous_value": 220 * 1024**3,
                "growth_ratio": 0.09,
                "window_start_at": "2026-05-22T13:00:00+00:00",
                "window_end_at": "2026-06-06T19:00:00+00:00",
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            content, _, _, _ = build_report_docx(
                FakeReportService(window_vms=chinese_vm).latest_report(period_days=30),
                settings,
                period_days=30,
            )
            word_xml, _ = _docx_xml(content)
            self.assertIn("一  摘要", word_xml)
            self.assertIn("关键发现", word_xml)
            self.assertIn("边缘云管理平台04", word_xml)
            self.assertIn("20.00 GB", word_xml)
            self.assertIn('<w:t xml:space="preserve"> 边缘云管理平台04 </w:t>', word_xml)
            self.assertIn('<w:t>Tower A</w:t>', word_xml)
            self.assertIn('<w:t>Cluster A</w:t>', word_xml)
            self.assertNotIn('<w:t xml:space="preserve">  20.00 GB  </w:t>', word_xml)
            self.assertIn('w:b/>', word_xml)
            self.assertIn('w:sz w:val="23"', word_xml)
            self.assertIn("全集群汇总报告", word_xml)
            self.assertIn("容量风险评估矩阵", word_xml)
            self.assertNotIn("单 VM 容量异常", word_xml)
            self.assertIn("重点 VM 容量", word_xml)
            self.assertIn("短期（本月）", word_xml)
            self.assertIn("中期（1-3 个月）", word_xml)
            self.assertIn("三个月以上", word_xml)
            self.assertIn("建议确认", word_xml)
            self.assertIn("启动扩容评估", word_xml)
            self.assertIn("基数较小", word_xml)

    def test_docx_vm_top_uses_actual_sample_window_not_requested_period(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_docx, build_report_xlsx

        month_vms = [
            {
                "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A", "vm_id": "vm-1", "vm": "VM One"},
                "forecast": {"status": "ok", "slope_per_day": 4, "current": 240 * 1024**3},
                "growth_amount": 150 * 1024**3,
                "previous_value": 120 * 1024**3,
                "growth_ratio": 1.0,
                "window_start_at": "2026-06-05T16:00:00+00:00",
                "window_end_at": "2026-06-06T02:00:00+00:00",
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-docx-secret")
            report = FakeReportService(month_vms=month_vms, window_vms=[]).latest_report(period_days=365)
            content, _, _, _ = build_report_docx(report, settings, period_days=365)
            word_xml, _ = _docx_xml(content)
            self.assertIn("统计窗口：2026-06-06 - 2026-06-06", word_xml)
            self.assertNotIn("统计窗口：2025-06-06 - 2026-06-06", word_xml)

            xlsx_content, _, _, _ = build_report_xlsx(report, settings, period_days=365)
            workbook_text = "\n".join(_xlsx_values(xlsx_content))
            self.assertIn("2026-06-06 - 2026-06-06", workbook_text)

    def test_xlsx_uses_customer_template_with_v2_scope_and_units(self) -> None:
        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_xlsx
        from openpyxl import load_workbook

        clusters = [
            {
                "labels": {"tower_id": "1", "tower": "Tower A", "cluster_id": "cluster-a", "cluster": "Cluster A"},
                "forecast": {
                    "status": "ok",
                    "slope_per_day": 10 * 1024**3,
                    "current": 810 * 1024**3,
                    "forecast_90d": 1090 * 1024**3,
                    "exhaustion_days": 64,
                },
                "points": [[1764547200, 780 * 1024**3], [1764720000, 810 * 1024**3]],
                "total": 1000 * 1024**3,
                "warning": 900 * 1024**3,
            },
            {
                "labels": {"tower_id": "2", "tower": "Tower B", "cluster_id": "cluster-b", "cluster": "Cluster B"},
                "forecast": {
                    "status": "ok",
                    "slope_per_day": 3 * 1024**3,
                    "current": 300 * 1024**3,
                    "forecast_90d": 570 * 1024**3,
                    "exhaustion_days": None,
                },
                "points": [[1764547200, 260 * 1024**3], [1764720000, 300 * 1024**3]],
                "total": 1000 * 1024**3,
                "warning": 900 * 1024**3,
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-xlsx-secret")
            report = FakeReportService(clusters=clusters).latest_report(period_days=14)
            content, _, _, _ = build_report_xlsx(report, settings, period_days=14)
            workbook = load_workbook(io.BytesIO(content), data_only=True)

            self.assertIn("封面", workbook.sheetnames)
            self.assertIn("执行摘要", workbook.sheetnames)
            self.assertIn("数据质量说明", workbook.sheetnames)
            self.assertIn("容量趋势", workbook.sheetnames)
            self.assertIn("VM增长TOP100", workbook.sheetnames)
            self.assertIn("日增长详情", workbook.sheetnames)
            self.assertIn("月增长详情", workbook.sheetnames)
            self.assertIn("本日新建VM", workbook.sheetnames)
            self.assertIn("本月新建VM", workbook.sheetnames)
            self.assertIn("Cluster A", workbook.sheetnames)
            self.assertIn("Cluster B", workbook.sheetnames)
            for removed_sheet in ["目录", "范围明细", "集群汇总", "VM_TOP100_汇总", "VM增长TOP20", "汇总"]:
                self.assertNotIn(removed_sheet, workbook.sheetnames)

            values = [str(cell) for sheet in workbook.worksheets for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None]
            workbook_text = "\n".join(values)
            self.assertIn("存储容量预测平台", workbook_text)
            self.assertIn("SMARTX超融合存储容量分析报告", workbook_text)
            self.assertIn("SMARTX HCI Storage Capacity Analysis Report", workbook_text)
            self.assertIn("客户名称：", workbook_text)
            self.assertIn("Tower范围：全部 Tower（2 个）", workbook_text)
            self.assertIn("集群范围：全部集群（2 个）", workbook_text)
            self.assertIn("近 14 天样本增长", workbook_text)
            self.assertIn("近两周增长 VM", workbook_text)
            self.assertIn("数据质量说明", workbook_text)
            self.assertIn("总体状态", workbook_text)
            self.assertIn("SQLite 当前 VM 数", workbook_text)
            self.assertIn("Prometheus 当前 VM series 数", workbook_text)
            self.assertIn("差异比例", workbook_text)
            self.assertIn("2026-06-02", workbook_text)
            self.assertIn("prometheus_cluster_sample_missing", workbook_text)
            self.assertIn("810.00 GiB", workbook_text)
            self.assertIn("1.06 TiB", workbook_text)
            self.assertIn("64 天", workbook_text)
            self.assertNotIn("SmartX 超融合平台", workbook_text)
            self.assertNotIn("CHINATOWER", workbook_text)
            self.assertNotIn("SMARTX-TT-WW", workbook_text)
            self.assertNotIn("22.27 TB", workbook_text)

            trend_headers = [cell.value for cell in workbook["容量趋势"][3]]
            self.assertIn("Tower", trend_headers)
            self.assertIn("集群", trend_headers)

            top_sheet = workbook["VM增长TOP100"]
            self.assertIn("TOP100", str(top_sheet["A1"].value))
            self.assertIn("TOP100", str(top_sheet["I1"].value))
            self.assertGreaterEqual(top_sheet.column_dimensions["B"].width, 36)
            self.assertGreaterEqual(top_sheet.column_dimensions["C"].width, 16)
            self.assertEqual(top_sheet.column_dimensions["E"].width, 13)
            self.assertIn(f"H1:H{top_sheet.max_row}", {str(item) for item in top_sheet.merged_cells.ranges})
            self.assertIsNone(top_sheet["H1"].fill.fill_type)
            self.assertEqual(top_sheet.column_dimensions["H"].width, 10)

            cluster_sheet = workbook["Cluster A"]
            cluster_values = [cell for row in cluster_sheet.iter_rows(values_only=True) for cell in row if cell is not None]
            cluster_text = "\n".join(str(value) for value in cluster_values)
            self.assertIn("Tower", cluster_text)
            self.assertIn("Cluster A", cluster_text)
            self.assertIn("810.00 GiB", cluster_text)
            self.assertIn("+30.00 GiB", cluster_text)
            self.assertIn("64 天", cluster_text)
            self.assertIn("增长量 TOP100", cluster_text)
            self.assertIn("增长率 TOP100", cluster_text)
            self.assertGreaterEqual(cluster_sheet.column_dimensions["A"].width, 36)
            self.assertGreaterEqual(cluster_sheet.column_dimensions["B"].width, 16)
            self.assertGreaterEqual(cluster_sheet.column_dimensions["D"].width, 16)
            self.assertEqual(
                [cluster_sheet.cell(row=2, column=column).value for column in range(1, 10)],
                ["Tower", "集群", "当前容量", "近 14 天样本增长", "风险", "90 天预测容量", "总容量", "使用率", "预计耗尽天数"],
            )
            self.assertEqual(cluster_sheet["A3"].value, "Tower A")
            self.assertEqual(cluster_sheet["B3"].value, "Cluster A")
            self.assertEqual(cluster_sheet["C3"].value, "810.00 GiB")
            self.assertEqual(cluster_sheet["I3"].value, "64 天")
            vm_header_rows = [
                row_index
                for row_index in range(1, cluster_sheet.max_row + 1)
                if cluster_sheet.cell(row=row_index, column=1).value == "VM"
                and cluster_sheet.cell(row=row_index, column=2).value == "当前容量"
            ]
            self.assertEqual(len(vm_header_rows), 2)
            for header_row in vm_header_rows:
                self.assertNotEqual(cluster_sheet.cell(row=header_row + 1, column=1).value, "VM")

            cover = workbook["封面"]
            self.assertEqual(cover["B8"].value, "SMARTX超融合存储容量分析报告")
            self.assertEqual(cover["B9"].value, "SMARTX HCI Storage Capacity Analysis Report")
            self.assertEqual(cover["B8"].font.color.rgb[-6:], "1A3C6E")
            self.assertEqual(cover["B8"].font.name, "Noto Sans CJK SC")
            self.assertEqual(cover["B8"].font.sz, 26)
            self.assertEqual(cover.row_dimensions[8].height, 40)

            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        self.assertEqual(cell.font.name, "Noto Sans CJK SC", f"{sheet.title}!{cell.coordinate}")
                        self.assertGreater(cell.font.sz or 0, 0, f"{sheet.title}!{cell.coordinate}")

    def test_xlsx_matches_optimized_customer_template_layout(self) -> None:
        from openpyxl import load_workbook

        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_xlsx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-xlsx-layout-secret")
            report = FakeReportService().latest_report(period_days=14)
            content, _, _, _ = build_report_xlsx(report, settings, period_days=14)
            workbook = load_workbook(io.BytesIO(content), data_only=True)

        self.assertEqual(
            workbook.sheetnames,
            [
                "封面",
                "执行摘要",
                "数据质量说明",
                "容量趋势",
                "VM增长TOP100",
                "日增长详情",
                "月增长详情",
                "本日新建VM",
                "本月新建VM",
                "Cluster A",
                "Cluster B",
            ],
        )

        cover = workbook["封面"]
        self.assertEqual(cover.row_dimensions[8].height, 40)
        self.assertEqual(cover.column_dimensions["A"].width, 8)
        self.assertEqual(cover.column_dimensions["B"].width, 50)
        self.assertEqual(cover.column_dimensions["C"].width, 8)
        self.assertEqual(cover.column_dimensions["D"].width, 20)
        self.assertEqual(cover["B8"].font.sz, 26)

        summary = workbook["执行摘要"]
        self.assertEqual(summary.column_dimensions["A"].width, 50)
        self.assertEqual(summary.column_dimensions["B"].width, 20.5)
        self.assertEqual(summary.column_dimensions["C"].width, 18)
        self.assertAlmostEqual(summary.column_dimensions["D"].width, 38.83203125)
        self.assertEqual(summary.row_dimensions[2].height, 38)
        self.assertEqual(summary["A5"].font.sz, 18)

        quality = workbook["数据质量说明"]
        self.assertEqual(quality.freeze_panes, "A4")
        self.assertEqual(quality["A1"].value, "数据质量说明")
        self.assertGreaterEqual(quality.column_dimensions["A"].width, 24)
        self.assertGreaterEqual(quality.column_dimensions["B"].width, 44)

        trend = workbook["容量趋势"]
        self.assertEqual(trend.freeze_panes, "A4")
        self.assertEqual(trend.column_dimensions["A"].width, 50)
        self.assertAlmostEqual(trend.column_dimensions["B"].width, 25.83203125)
        self.assertAlmostEqual(trend.column_dimensions["C"].width, 16.83203125)
        self.assertEqual(trend.row_dimensions[2].height, 51)
        self.assertEqual(trend.row_dimensions[4].height, 30)

        top = workbook["VM增长TOP100"]
        self.assertEqual(top.freeze_panes, "A5")
        self.assertEqual(top.sheet_view.topLeftCell, "A1")
        self.assertEqual(top.sheet_view.selection[0].activeCell, "A5")
        self.assertEqual(top.sheet_view.selection[0].sqref, "A5")
        self.assertAlmostEqual(top.column_dimensions["A"].width, 15.83203125)
        self.assertEqual(top.column_dimensions["B"].width, 50)
        self.assertEqual(top.column_dimensions["H"].width, 10)
        self.assertAlmostEqual(top.column_dimensions["I"].width, 15.83203125)
        self.assertEqual(top.column_dimensions["J"].width, 50)
        self.assertEqual(top.row_dimensions[1].height, 34)
        self.assertEqual(top.row_dimensions[3].height, 34)
        self.assertEqual(top.row_dimensions[4].height, 28)
        self.assertIn(f"H1:H{top.max_row}", {str(item) for item in top.merged_cells.ranges})
        self.assertEqual(top["A4"].font.sz, 11)
        self.assertEqual(top["A5"].font.sz, 11)

        day_growth = workbook["日增长详情"]
        self.assertEqual(day_growth.column_dimensions["A"].width, 15.83203125)
        self.assertEqual(day_growth.column_dimensions["B"].width, 44.5)
        self.assertAlmostEqual(day_growth.column_dimensions["C"].width, 25.83203125)
        self.assertEqual(day_growth.row_dimensions[4].height, 30)

        day_new = workbook["本日新建VM"]
        self.assertEqual(day_new.column_dimensions["A"].width, 17.5)
        self.assertEqual(day_new.column_dimensions["B"].width, 18)
        self.assertEqual(day_new.column_dimensions["C"].width, 36)
        self.assertEqual(day_new.row_dimensions[1].height, 22)

        cluster = workbook["Cluster A"]
        self.assertAlmostEqual(cluster.column_dimensions["A"].width, 36.83203125)
        self.assertEqual(cluster.column_dimensions["B"].width, 39.5)
        self.assertAlmostEqual(cluster.column_dimensions["C"].width, 22.6640625)
        self.assertAlmostEqual(cluster.column_dimensions["D"].width, 19.6640625)
        self.assertEqual(cluster.column_dimensions["E"].width, 15)
        self.assertEqual(cluster.row_dimensions[1].height, 32)
        self.assertEqual(cluster.row_dimensions[2].height, 24)
        self.assertEqual(cluster.row_dimensions[3].height, 24)
        self.assertEqual(cluster.row_dimensions[4].height, 32)
        self.assertEqual(cluster["A1"].font.sz, 23)
        self.assertEqual(cluster["A3"].font.sz, 12)
        self.assertEqual(cluster["A6"].font.sz, 14)

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        self.assertEqual(cell.font.name, "Noto Sans CJK SC", f"{sheet.title}!{cell.coordinate}")

    def test_xlsx_summary_and_growth_detail_sheets_follow_customer_layout(self) -> None:
        from openpyxl import load_workbook

        from app.v2.config import V2Settings
        from app.v2.reports.export import build_report_xlsx

        with tempfile.TemporaryDirectory() as tmpdir:
            settings = V2Settings(data_root=Path(tmpdir), secret_key="reports-xlsx-growth-detail-secret")
            report = FakeReportService().latest_report(period_days=14)
            content, _, _, _ = build_report_xlsx(report, settings, period_days=14)
            workbook = load_workbook(io.BytesIO(content), data_only=True)

        summary = workbook["执行摘要"]
        for row_index in [4, 6]:
            for column_index in range(1, 5):
                cell = summary.cell(row=row_index, column=column_index)
                self.assertTrue(cell.font.bold, cell.coordinate)
                self.assertEqual(cell.font.color.type, "rgb", cell.coordinate)
                self.assertEqual(cell.font.color.rgb[-6:], "000000", cell.coordinate)

        day_sheet = workbook["日增长详情"]
        self.assertIn("A1:I1", {str(item) for item in day_sheet.merged_cells.ranges})
        self.assertIn("日增长最快虚拟机", str(day_sheet["A1"].value))
        self.assertIsNone(day_sheet.sheet_properties.tabColor)

        self.assertEqual(
            workbook.sheetnames[5:8],
            ["日增长详情", "月增长详情", "本日新建VM"],
        )
        month_sheet = workbook["月增长详情"]
        self.assertIn("A1:I1", {str(item) for item in month_sheet.merged_cells.ranges})
        self.assertIn("月增长最快虚拟机", str(month_sheet["A1"].value))
        self.assertEqual(month_sheet["G3"].value, "月增长量")
        month_text = "\n".join(
            str(cell)
            for row in month_sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        )
        self.assertIn("VM One", month_text)
        self.assertNotIn("Day VM", month_text)


@unittest.skipIf(TestClient is None, "FastAPI test dependencies are not installed.")
class V2ReportExportApiTest(unittest.TestCase):
    def test_report_exports_require_auth_save_files_and_expose_download_link(self) -> None:
        from app.v2.api import get_report_service
        from app.v2.main import create_app

        with tempfile.TemporaryDirectory() as tmpdir:
            os.environ["SMARTX_DATA_ROOT"] = tmpdir
            os.environ["SMARTX_SECRET_KEY"] = "reports-export-secret"
            os.environ["SMARTX_ADMIN_PASSWORD"] = "password"
            try:
                app = create_app()
                app.dependency_overrides[get_report_service] = lambda: FakeReportService()
                with TestClient(app) as client:
                    self.assertEqual(client.get("/api/reports/export/word").status_code, 401)
                    token = client.post("/api/auth/login", json={"username": "admin", "password": "password"}).json()["access_token"]
                    headers = {"Authorization": f"Bearer {token}"}

                    word = client.get("/api/reports/export/word?period_days=30", headers=headers)
                    self.assertEqual(word.status_code, 200)
                    self.assertEqual(word.headers["content-type"], "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
                    self.assertIn("storage-forecast-all-", word.headers["content-disposition"])
                    word_path = Path(word.headers["x-smartx-export-path"])
                    self.assertTrue(word_path.is_file())
                    self.assertEqual(word_path.parent, Path(tmpdir) / "exports" / "reports")
                    self.assertTrue(word.headers["x-smartx-export-url"].startswith("/api/admin/exports/reports/"))
                    self.assertGreater(len(word.content), 1000)
                    word_xml, _ = _docx_xml(word.content)
                    self.assertIn("一  摘要", word_xml)
                    self.assertIn("目录", word_xml)
                    self.assertIn("风险状态", word_xml)
                    self.assertIn("高风险", word_xml)
                    self.assertIn("Cluster A", word_xml)
                    self.assertIn("Cluster B", word_xml)
                    self.assertIn('w:fill="1A3C6E"', word_xml)
                    self.assertIn('w:fill="F0F4FA"', word_xml)
                    self.assertIn('w:color w:val="007ACC"', word_xml)

                    excel = client.get("/api/reports/export/excel?period_days=30", headers=headers)
                    self.assertEqual(excel.status_code, 200)
                    self.assertEqual(excel.headers["content-type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    excel_path = Path(excel.headers["x-smartx-export-path"])
                    self.assertTrue(excel_path.is_file())
                    self.assertEqual(excel_path.parent, Path(tmpdir) / "exports" / "reports")
                    self.assertGreater(len(excel.content), 1000)
                    with zipfile.ZipFile(excel_path) as workbook:
                        styles_xml = workbook.read("xl/styles.xml").decode("utf-8")
                    from openpyxl import load_workbook

                    workbook = load_workbook(excel_path, read_only=True)
                    self.assertIn("执行摘要", workbook.sheetnames)
                    self.assertIn("VM增长TOP100", workbook.sheetnames)
                    self.assertIn("Cluster A", workbook.sheetnames)
                    for removed_sheet in ["目录", "范围明细", "集群汇总", "VM_TOP100_汇总", "VM增长TOP20"]:
                        self.assertNotIn(removed_sheet, workbook.sheetnames)
                    summary_values = [cell for row in workbook["执行摘要"].iter_rows(values_only=True) for cell in row if cell]
                    self.assertIn("容量风险摘要", summary_values)
                    self.assertTrue(any("Cluster A 使用率超过 80%" in str(value) for value in summary_values))
                    self.assertIn("F4CCCC", styles_xml)

                    download = client.get(word.headers["x-smartx-export-url"], headers=headers)
                    self.assertEqual(download.status_code, 200)
                    self.assertEqual(download.content, word.content)
                    word_path.unlink()
                    expired_download = client.get(word.headers["x-smartx-export-url"], headers=headers)
                    self.assertEqual(expired_download.status_code, 404)
                    self.assertIn("已失效", expired_download.json()["detail"])

                    tasks = client.get("/api/tasks", headers=headers).json()
                    report_tasks = [task for task in tasks if task["type"] == "report"]
                    self.assertGreaterEqual(len(report_tasks), 2)
                    self.assertEqual(report_tasks[0]["status"], "success")
                    self.assertTrue(report_tasks[0]["links"][0]["url"].startswith("/api/admin/exports/reports/"))
                    expired_links = [link for task in report_tasks for link in task["links"] if link.get("filename") == word_path.name]
                    self.assertTrue(any(link.get("expired") for link in expired_links))
            finally:
                os.environ.pop("SMARTX_DATA_ROOT", None)
                os.environ.pop("SMARTX_SECRET_KEY", None)
                os.environ.pop("SMARTX_ADMIN_PASSWORD", None)

    def test_report_bundle_exports_word_and_excel_with_one_task(self) -> None:
        from app.v2.api import get_report_service
        from app.v2.main import create_app
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            os.environ["SMARTX_DATA_ROOT"] = tmpdir
            os.environ["SMARTX_SECRET_KEY"] = "reports-bundle-secret"
            os.environ["SMARTX_ADMIN_PASSWORD"] = "password"
            try:
                app = create_app()
                app.dependency_overrides[get_report_service] = lambda: FakeReportService(month_vms=[])
                with TestClient(app) as client:
                    self.assertEqual(client.post("/api/reports/export/bundle?period_days=30").status_code, 401)
                    token = client.post("/api/auth/login", json={"username": "admin", "password": "password"}).json()["access_token"]
                    headers = {"Authorization": f"Bearer {token}"}

                    response = client.post("/api/reports/export/bundle?period_days=30&task_id=report-export-ui-1", headers=headers)
                    self.assertEqual(response.status_code, 200)
                    payload = response.json()
                    self.assertEqual(payload["task_id"], "report-export-ui-1")
                    self.assertEqual(payload["status"], "success")
                    self.assertEqual(payload["message"], "Word 和 Excel 报表已生成")
                    self.assertEqual(len(payload["files"]), 2)
                    self.assertEqual([link["label"] for link in payload["links"]], ["Word", "Excel"])

                    word_file = next(file for file in payload["files"] if file["label"] == "Word")
                    excel_file = next(file for file in payload["files"] if file["label"] == "Excel")
                    word_path = Path(word_file["path"])
                    excel_path = Path(excel_file["path"])
                    self.assertTrue(word_path.is_file())
                    self.assertTrue(excel_path.is_file())
                    self.assertEqual(word_path.parent, Path(tmpdir) / "exports" / "reports")
                    self.assertEqual(excel_path.parent, Path(tmpdir) / "exports" / "reports")

                    word_xml, _ = _docx_xml(word_path.read_bytes())
                    self.assertIn("存储容量预测平台", word_xml)
                    self.assertIn("SMARTX超融合存储容量分析报告", word_xml)
                    self.assertIn("SMARTX HCI Storage Capacity Analysis Report", word_xml)
                    self.assertNotIn("SmartX 超融合平台", word_xml)
                    self.assertNotIn("Storage Capacity Forecast Report", word_xml)
                    self.assertIn("目录", word_xml)
                    self.assertIn("一  摘要", word_xml)
                    self.assertIn("二  集群容量概览", word_xml)
                    self.assertIn("三  集群虚拟机增长分析", word_xml)
                    self.assertIn("四  集群容量趋势图表", word_xml)
                    self.assertIn("五  全集群汇总报告", word_xml)
                    self.assertIn("2.1  范围基本信息", word_xml)
                    self.assertIn("3.1  Tower A / Cluster A", word_xml)
                    self.assertIn("3.2  Tower A / Cluster B", word_xml)
                    self.assertIn("4.1  Tower A / Cluster A", word_xml)
                    self.assertIn("4.2  Tower A / Cluster B", word_xml)
                    self.assertIn("Top 10 VM 增长量", word_xml)
                    self.assertIn("增长 VM 样本数", word_xml)
                    self.assertNotIn("虚拟机所属人", word_xml)

                    workbook = load_workbook(excel_path, read_only=True)
                    self.assertIn("封面", workbook.sheetnames)
                    self.assertIn("执行摘要", workbook.sheetnames)
                    self.assertIn("VM增长TOP100", workbook.sheetnames)
                    for removed_sheet in ["目录", "范围明细", "集群汇总", "VM_TOP100_汇总", "VM增长TOP20"]:
                        self.assertNotIn(removed_sheet, workbook.sheetnames)
                    self.assertTrue(any(name.startswith("Cluster A") for name in workbook.sheetnames))
                    workbook_values = [
                        str(cell)
                        for sheet in workbook.worksheets
                        for row in sheet.iter_rows(values_only=True)
                        for cell in row
                        if cell is not None
                    ]
                    self.assertNotIn("虚拟机所属人", workbook_values)
                    self.assertIn("当前 30 天窗口暂无明显 VM 增长数据", workbook_values)

                    tasks = client.get("/api/tasks", headers=headers).json()
                    report_tasks = [task for task in tasks if task["type"] == "report"]
                    self.assertEqual(len(report_tasks), 1)
                    self.assertEqual(report_tasks[0]["title"], "导出预测报表")
                    self.assertEqual(report_tasks[0]["status"], "success")
                    self.assertEqual([link["label"] for link in report_tasks[0]["links"]], ["Word", "Excel"])
            finally:
                os.environ.pop("SMARTX_DATA_ROOT", None)
                os.environ.pop("SMARTX_SECRET_KEY", None)
                os.environ.pop("SMARTX_ADMIN_PASSWORD", None)


if __name__ == "__main__":
    unittest.main()


def _docx_xml(content: bytes) -> tuple[str, str]:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        document_xml = archive.read("word/document.xml").decode("utf-8")
        footer_xml = "\n".join(
            archive.read(name).decode("utf-8")
            for name in archive.namelist()
            if name.startswith("word/footer")
        )
    return document_xml, footer_xml


def _docx_paragraph_containing(document_xml: str, text: str) -> str:
    for paragraph in document_xml.split("</w:p>"):
        if text in paragraph:
            return paragraph
    raise AssertionError(f"DOCX paragraph not found: {text}")


def _docx_row_containing(document_xml: str, text: str) -> str:
    for row in document_xml.split("<w:tr")[1:]:
        row_xml = "<w:tr" + row.split("</w:tr>", 1)[0] + "</w:tr>"
        if text in row_xml:
            return row_xml
    raise AssertionError(f"DOCX table row not found: {text}")


def _xlsx_values(content: bytes) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return [
        str(cell)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    ]
